"""
MS_recording&maintenance — 质谱仪使用记录与维护管理
支持 Q-IM-TOF / LTQ / Q-Exactive 三种质谱仪
"""

import os
import re
import sys
import csv
import json
import sqlite3
import subprocess
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, timedelta

# ============================================================
# 路径配置
# ============================================================
if getattr(sys, "frozen", False):
    APP_DIR = os.path.dirname(sys.executable)
else:
    APP_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(APP_DIR, "ms_data.db")
CONFIG_PATH = os.path.join(APP_DIR, "ms_config.json")

# ============================================================
# 静态数据
# ============================================================
ATOMIC_MASSES = {
    "H": 1.00794, "He": 4.002602, "Li": 6.941, "Be": 9.012182, "B": 10.811,
    "C": 12.0107, "N": 14.0067, "O": 15.9994, "F": 18.9984032, "Ne": 20.1797,
    "Na": 22.98977, "Mg": 24.305, "Al": 26.98154, "Si": 28.0855, "P": 30.97376,
    "S": 32.065, "Cl": 35.453, "Ar": 39.948, "K": 39.0983, "Ca": 40.078,
    "Sc": 44.9559, "Ti": 47.867, "V": 50.9415, "Cr": 51.9961, "Mn": 54.938,
    "Fe": 55.845, "Co": 58.9332, "Ni": 58.6934, "Cu": 63.546, "Zn": 65.39,
    "Ga": 69.723, "Ge": 72.64, "As": 74.9216, "Se": 78.96, "Br": 79.904,
    "Kr": 83.8, "Rb": 85.4678, "Sr": 87.62, "Y": 88.9059, "Zr": 91.224,
    "Nb": 92.9064, "Mo": 95.94, "Tc": 98.0, "Ru": 101.07, "Rh": 102.9055,
    "Pd": 106.42, "Ag": 107.8682, "Cd": 112.411, "In": 114.818, "Sn": 118.71,
    "Sb": 121.76, "Te": 127.6, "I": 126.9045, "Xe": 131.293, "Cs": 132.9055,
    "Ba": 137.327, "La": 138.9055, "Ce": 140.116, "Pr": 140.9077, "Nd": 144.24,
    "Pm": 145.0, "Sm": 150.36, "Eu": 151.964, "Gd": 157.25, "Tb": 158.9253,
    "Dy": 162.5, "Ho": 164.9303, "Er": 167.259, "Tm": 168.9342, "Yb": 173.04,
    "Lu": 174.967, "Hf": 178.49, "Ta": 180.9479, "W": 183.84, "Re": 186.207,
    "Os": 190.23, "Ir": 192.217, "Pt": 195.078, "Au": 196.9665, "Hg": 200.59,
    "Tl": 204.3833, "Pb": 207.2, "Bi": 208.9804, "Po": 209.0, "At": 210.0,
    "Rn": 222.0, "Fr": 223.0, "Ra": 226.0, "Ac": 227.0, "Th": 232.0381,
    "Pa": 231.0359, "U": 238.0289, "Np": 237.0, "Pu": 244.0, "Am": 243.0,
    "Cm": 247.0, "Bk": 247.0, "Cf": 251.0, "Es": 252.0, "Fm": 257.0,
    "Md": 258.0, "No": 259.0, "Lr": 262.0, "Rf": 261.0, "Db": 262.0,
    "Sg": 266.0, "Bh": 264.0, "Hs": 277.0, "Mt": 268.0,
}

MS_TYPES = ["Q-IM-TOF", "LTQ", "Q-Exactive"]

MS_CONFIG = {
    "Q-IM-TOF": {
        "ion_label": "测试条件",
        "ion_default": "2.50-40-80-100-20-50-600-4.6",
    },
    "LTQ": {
        "ion_label": "测试条件",
        "ion_default": "5kV-275℃ 10-50V",
    },
    "Q-Exactive": {
        "ion_label": "测试条件",
        "ion_default": "",
    },
}

# 每种质谱的维护类型及推荐周期(天)
# 根据维护计划表格 maintenance.xlsx 设定
MAINTENANCE_TYPES = {
    "Q-IM-TOF": [
        ("Q-IM-TOF ESI喷针+进样管路清洗", 30),      # 一个月
        ("Q-IM-TOF 质量校准与评估", 30),            # 一个月
        ("Q-IM-TOF 钢瓶余量", 30),                  # 一个月
        ("Q-IM-TOF StepWave", 180),                 # 半年
    ],
    "LTQ": [
        ("LTQ ESI喷针+进样管路清洗", 30),            # 一个月
        ("LTQ 离子传输管的清洗", 30),                # 一个月
        ("LTQ 机械泵震气", 7),                      # 一周
        ("LTQ 机械泵油量", 7),                       # 一周
        ("LTQ 更换泵油", 90),                       # 每季度/半年（推荐90天）
        ("LTQ 质量校准与评估", 30),                  # 一个月
        ("LTQ 钢瓶余量", 30),                        # 一个月
        ("LTQ API Stack的清洗", 60),                 # 两个月
    ],
    "Q-Exactive": [
        ("Q-Exactive ESI喷针+进样管路清洗", 30),      # 一个月
        ("Q-Exactive 离子传输管的清洗", 30),           # 一个月
        ("Q-Exactive 机械泵震气", 7),                 # 一周
        ("Q-Exactive 机械泵油量", 7),                  # 一周
        ("Q-Exactive 更换泵油", 90),                  # 每季度/半年（推荐90天）
        ("Q-Exactive 质量校准与评估", 30),             # 一个月
        ("Q-Exactive 钢瓶余量", 30),                  # 一个月
        ("Q-Exactive API Stack的清洗", 60),            # 两个月
    ],
}

# ============================================================
# 深色主题配色常量
# ============================================================
BG_DARK = "#0d1117"
BG_PANEL = "#161b22"
BG_INPUT = "#21262d"
BG_TREE = "#161b22"
BG_TREE_ALT = "#1c2333"
FG_PRIMARY = "#e6edf3"
FG_SECONDARY = "#8b949e"
ACCENT = "#58a6ff"
ACCENT_HIGHLIGHT = "#79c0ff"
GREEN_STATUS = "#3fb950"
YELLOW_STATUS = "#d2991d"
RED_STATUS = "#f85149"
BORDER_COLOR = "#30363d"

STATUS_COLORS = {
    "正常": GREEN_STATUS,
    "需关注": YELLOW_STATUS,
    "异常": RED_STATUS,
}

# ============================================================
# 工具函数
# ============================================================
def parse_number(s, i):
    if i >= len(s) or not s[i].isdigit():
        return 1, i
    num = 0
    while i < len(s) and s[i].isdigit():
        num = num * 10 + int(s[i])
        i += 1
    return num, i

def parse_formula(formula, i=0):
    total = 0.0
    while i < len(formula):
        c = formula[i]
        if c in "([{":
            sub_mass, i = parse_formula(formula, i + 1)
            if i >= len(formula) or formula[i] not in ")]}":
                raise ValueError(f"括号不匹配，位置 {i}")
            i += 1
            num, i = parse_number(formula, i)
            total += sub_mass * num
        elif c in ")]}":
            return total, i
        elif c in "·.":
            i += 1
        elif c.isupper():
            elem = c
            i += 1
            while i < len(formula) and formula[i].islower():
                elem += formula[i]
                i += 1
            num, i = parse_number(formula, i)
            mass = ATOMIC_MASSES.get(elem)
            if mass is None:
                raise ValueError(f"未知元素: {elem}")
            total += mass * num
        else:
            i += 1
    return total, i

def calc_mz(sample_info, charge_str):
    mass = parse_formula(sample_info)[0]
    charge_str = charge_str.strip()
    if not charge_str:
        return None, "请填写电荷信息"
    s = charge_str
    sign = 1
    if s.endswith("+") or s.endswith("-"):
        if s.endswith("-"):
            sign = -1
        num_str = s[:-1]
    elif s.startswith("+") or s.startswith("-"):
        if s.startswith("-"):
            sign = -1
        num_str = s[1:]
    else:
        return None, f"无法解析电荷: {charge_str}"
    z = 1 if num_str == "" else int(num_str)
    return mass / z, None

# ============================================================
# 配置读写
# ============================================================
def load_config():
    try:
        with open(CONFIG_PATH) as f:
            cfg = json.load(f)
        # 兼容旧配置：Q-Orbitrap → Q-Exactive 重命名
        if cfg.get("ms_type") == "Q-Orbitrap":
            cfg["ms_type"] = "Q-Exactive"
            save_config(cfg)
        return cfg
    except (FileNotFoundError, json.JSONDecodeError):
        return {}

def save_config(cfg):
    with open(CONFIG_PATH, "w") as f:
        json.dump(cfg, f, indent=2)

# ============================================================
# 数据库
# ============================================================
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn

def init_db():
    with get_db() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS experiments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT NOT NULL DEFAULT '',
                time_period TEXT NOT NULL DEFAULT '',
                purpose TEXT NOT NULL DEFAULT '',
                name TEXT NOT NULL DEFAULT '',
                ion_source TEXT NOT NULL DEFAULT '',
                sample_info TEXT NOT NULL DEFAULT '',
                charge TEXT NOT NULL DEFAULT '',
                sample_peaks TEXT NOT NULL DEFAULT '',
                solvent TEXT NOT NULL DEFAULT '',
                cleaned TEXT NOT NULL DEFAULT '',
                ms_type TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL DEFAULT (datetime('now', 'localtime'))
            )
        """)
        for col in [
            ("charge", "TEXT NOT NULL DEFAULT ''"),
            ("time_period", "TEXT NOT NULL DEFAULT ''"),
            ("solvent", "TEXT NOT NULL DEFAULT ''"),
            ("cleaned", "TEXT NOT NULL DEFAULT ''"),
            ("ms_type", "TEXT NOT NULL DEFAULT ''"),
        ]:
            try:
                conn.execute(f"ALTER TABLE experiments ADD COLUMN {col[0]} {col[1]}")
            except sqlite3.OperationalError:
                pass

        conn.execute("""
            CREATE TABLE IF NOT EXISTS maintenance (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ms_type TEXT NOT NULL DEFAULT '',
                date TEXT NOT NULL DEFAULT '',
                name TEXT NOT NULL DEFAULT '',
                record_type TEXT NOT NULL DEFAULT '',
                status TEXT NOT NULL DEFAULT '',
                notes TEXT NOT NULL DEFAULT '',
                next_date TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL DEFAULT (datetime('now', 'localtime'))
            )
        """)
        for col in [
            ("name", "TEXT NOT NULL DEFAULT ''"),
            ("status", "TEXT NOT NULL DEFAULT ''"),
            ("next_date", "TEXT NOT NULL DEFAULT ''"),
        ]:
            try:
                conn.execute(f"ALTER TABLE maintenance ADD COLUMN {col[0]} {col[1]}")
            except sqlite3.OperationalError:
                pass

# ============================================================
# 深色主题样式
# ============================================================
def apply_dark_theme(root):
    style = ttk.Style(root)
    style.theme_use("clam")

    root.configure(bg=BG_DARK)

    # 通用配置
    style.configure(".", background=BG_DARK, foreground=FG_PRIMARY,
                    fieldbackground=BG_INPUT, borderwidth=1,
                    relief="flat", font=("Segoe UI", 9))

    # TFrame
    style.configure("TFrame", background=BG_DARK)
    style.configure("Panel.TFrame", background=BG_PANEL)
    style.configure("InputPanel.TFrame", background=BG_PANEL)
    style.configure("Card.TFrame", background=BG_PANEL, borderwidth=1,
                    relief="solid", bordercolor=BORDER_COLOR)

    # TLabel
    style.configure("TLabel", background=BG_DARK, foreground=FG_PRIMARY,
                    font=("Segoe UI", 9))
    style.configure("Panel.TLabel", background=BG_PANEL, foreground=FG_PRIMARY)
    style.configure("Header.TLabel", background=BG_DARK, foreground=ACCENT_HIGHLIGHT,
                    font=("Segoe UI", 13, "bold"))
    style.configure("Title.TLabel", background=BG_DARK, foreground=ACCENT_HIGHLIGHT,
                    font=("Segoe UI", 12, "bold"))
    style.configure("Status.TLabel", background=BG_DARK, foreground=GREEN_STATUS,
                    font=("Segoe UI", 9, "bold"))
    style.configure("Warn.TLabel", background=BG_DARK, foreground=YELLOW_STATUS,
                    font=("Segoe UI", 9, "bold"))
    style.configure("Hint.TLabel", background=BG_DARK, foreground=FG_SECONDARY,
                    font=("Segoe UI", 8))

    # TButton
    style.configure("TButton", background=BG_INPUT, foreground=ACCENT,
                    borderwidth=1, bordercolor=ACCENT, padding=(14, 6),
                    font=("Segoe UI", 9, "bold"), relief="solid")
    style.map("TButton",
              background=[("active", "#30363d"), ("pressed", "#1c2638")],
              foreground=[("active", ACCENT_HIGHLIGHT)],
              bordercolor=[("active", ACCENT_HIGHLIGHT)])
    style.configure("Primary.TButton", background="#1f6feb", foreground="#ffffff",
                    borderwidth=1, bordercolor="#1f6feb", padding=(16, 6),
                    font=("Segoe UI", 9, "bold"))
    style.map("Primary.TButton",
              background=[("active", "#388bfd"), ("pressed", "#1f6feb")],
              foreground=[("active", "#ffffff")])
    style.configure("Danger.TButton", background="#da3633", foreground="#ffffff",
                    borderwidth=1, bordercolor="#da3633", padding=(14, 6),
                    font=("Segoe UI", 9, "bold"))
    style.map("Danger.TButton",
              background=[("active", "#f85149"), ("pressed", "#da3633")])

    # TEntry
    style.configure("TEntry", fieldbackground=BG_INPUT, foreground=FG_PRIMARY,
                    borderwidth=1, bordercolor=BORDER_COLOR, padding=(8, 5),
                    font=("Segoe UI", 9))
    style.map("TEntry",
              fieldbackground=[("focus", "#1c2638")],
              bordercolor=[("focus", ACCENT)])

    # TCombobox
    style.configure("TCombobox", fieldbackground=BG_INPUT, foreground=FG_PRIMARY,
                    borderwidth=1, bordercolor=BORDER_COLOR, padding=(8, 5),
                    arrowcolor=FG_PRIMARY, background=BG_INPUT)
    style.map("TCombobox",
              fieldbackground=[("readonly", BG_INPUT), ("focus", "#1c2638")],
              bordercolor=[("focus", ACCENT)])

    # 让下拉列表变深色
    root.option_add("*TCombobox*Listbox.background", BG_INPUT)
    root.option_add("*TCombobox*Listbox.foreground", FG_PRIMARY)
    root.option_add("*TCombobox*Listbox.selectBackground", "#388bfd")
    root.option_add("*TCombobox*Listbox.selectForeground", "#ffffff")
    root.option_add("*TCombobox*Listbox.font", ("Segoe UI", 9))

    # Treeview
    style.configure("Treeview", background=BG_TREE, foreground=FG_PRIMARY,
                    fieldbackground=BG_TREE, borderwidth=1, bordercolor=BORDER_COLOR,
                    rowheight=28, font=("Segoe UI", 9))
    style.configure("Treeview.Heading", background=BG_INPUT, foreground=FG_PRIMARY,
                    borderwidth=1, bordercolor=BORDER_COLOR, padding=(8, 6),
                    font=("Segoe UI", 9, "bold"), relief="flat")
    style.map("Treeview.Heading",
              background=[("active", "#30363d")])
    style.map("Treeview",
              background=[("selected", "#1f6feb")],
              foreground=[("selected", "#ffffff")])

    # TSeparator
    style.configure("TSeparator", background=BORDER_COLOR)

    # TScrollbar
    style.configure("TScrollbar", background=BG_INPUT, troughcolor=BG_DARK,
                    borderwidth=0, arrowcolor=FG_PRIMARY)

    return style


# ============================================================
# 质谱类型选择窗口
# ============================================================
class TypeSelector(tk.Toplevel):
    """启动时选择质谱类型"""
    def __init__(self):
        super().__init__()
        self.title("选择质谱类型")
        self.result = None
        self.configure(bg=BG_DARK)

        w, h = 560, 400
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        x = (sw - w) // 2
        y = (sh - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")
        self.resizable(False, False)

        # 顶部装饰线
        canvas = tk.Canvas(self, height=3, bg=BG_DARK, highlightthickness=0)
        canvas.pack(fill=tk.X)
        canvas.create_rectangle(0, 0, w, 3, fill=ACCENT, outline="")

        main = tk.Frame(self, bg=BG_DARK)
        main.pack(fill=tk.BOTH, expand=True, padx=40, pady=(30, 20))

        tk.Label(main, text="MS Recording & Maintenance",
                 font=("Segoe UI", 18, "bold"), fg=ACCENT_HIGHLIGHT, bg=BG_DARK).pack(pady=(0, 8))
        tk.Label(main, text="质谱仪使用记录与维护管理系统",
                 font=("Segoe UI", 10), fg=FG_SECONDARY, bg=BG_DARK).pack(pady=(0, 24))

        tk.Label(main, text="请选择质谱类型",
                 font=("Segoe UI", 12), fg=FG_PRIMARY, bg=BG_DARK).pack(pady=(0, 20))

        btn_frame = tk.Frame(main, bg=BG_DARK)
        btn_frame.pack()

        for i, ms_type in enumerate(MS_TYPES):
            btn = tk.Button(
                btn_frame, text=ms_type, width=22, height=2,
                font=("Segoe UI", 11, "bold"),
                bg=BG_INPUT, fg=ACCENT_HIGHLIGHT,
                activebackground="#30363d", activeforeground="#ffffff",
                relief="flat", bd=1, cursor="hand2",
                command=lambda t=ms_type: self._select(t),
            )
            btn.grid(row=i, column=0, pady=5)
            btn.bind("<Enter>", lambda e, b=btn: b.configure(bg="#1c2638"))
            btn.bind("<Leave>", lambda e, b=btn: b.configure(bg=BG_INPUT))

        self.protocol("WM_DELETE_WINDOW", self._on_close)

    def _select(self, ms_type):
        self.result = ms_type
        cfg = load_config()
        cfg["ms_type"] = ms_type
        save_config(cfg)
        self.destroy()

    def _on_close(self):
        self.result = None
        self.destroy()


# ============================================================
# 维护记录窗口
# ============================================================
class MaintenanceWindow(tk.Toplevel):
    """维护记录管理窗口"""
    def __init__(self, master, ms_type):
        super().__init__(master)
        self.ms_type = ms_type
        self.title(f"{ms_type} 维护记录")
        self.geometry("960x680")
        self.minsize(700, 500)
        self.configure(bg=BG_DARK)
        self.editing_id = None
        self._sort_dir = {}
        self._all_rows = []
        self._rclick_start = None

        # 维护类型映射
        self.mt_config = {t[0]: t[1] for t in MAINTENANCE_TYPES.get(ms_type, [])}
        self.mt_list = list(self.mt_config.keys())

        self._build_ui()
        self._load_data()

        # 窗口居中
        self.update_idletasks()
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        w = self.winfo_width()
        h = self.winfo_height()
        x = (sw - w) // 2
        y = (sh - h) // 2
        self.geometry(f"+{x}+{y}")

    def _build_ui(self):
        # ---- 顶部装饰 ----
        hdr = tk.Frame(self, bg=BG_DARK, height=40)
        hdr.pack(fill=tk.X, padx=16, pady=(16, 0))
        tk.Label(hdr, text=f"🔧 {self.ms_type} 维护记录",
                 font=("Segoe UI", 14, "bold"), fg=ACCENT_HIGHLIGHT, bg=BG_DARK).pack(side=tk.LEFT)

        tk.Label(hdr, text=f"维护项: {len(self.mt_list)} 项",
                 font=("Segoe UI", 9), fg=FG_SECONDARY, bg=BG_DARK).pack(side=tk.RIGHT)

        # ---- 输入区 ----
        panel = tk.Frame(self, bg=BG_PANEL, padx=16, pady=12)
        panel.pack(fill=tk.X, padx=16, pady=(8, 0))

        # 第一行：日期 + 维护类型 + 状态 + 下次维护
        r0 = tk.Frame(panel, bg=BG_PANEL)
        r0.pack(fill=tk.X, pady=(0, 4))

        # 日期
        tk.Label(r0, text="日期", fg=FG_PRIMARY, bg=BG_PANEL,
                 font=("Segoe UI", 9)).pack(side=tk.LEFT)
        self.var_date = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        self.ent_date = ttk.Entry(r0, textvariable=self.var_date, width=14)
        self.ent_date.pack(side=tk.LEFT, padx=(4, 16))

        # 姓名
        tk.Label(r0, text="姓名", fg=FG_PRIMARY, bg=BG_PANEL,
                 font=("Segoe UI", 9)).pack(side=tk.LEFT)
        self.var_name = tk.StringVar()
        self.ent_name = ttk.Entry(r0, textvariable=self.var_name, width=12)
        self.ent_name.pack(side=tk.LEFT, padx=(4, 16))

        # 维护类型
        tk.Label(r0, text="维护类型", fg=FG_PRIMARY, bg=BG_PANEL,
                 font=("Segoe UI", 9)).pack(side=tk.LEFT)
        self.var_record_type = tk.StringVar()
        self.cmb_type = ttk.Combobox(r0, textvariable=self.var_record_type,
                                      values=self.mt_list, width=24, state="readonly")
        self.cmb_type.pack(side=tk.LEFT, padx=4)
        self.cmb_type.bind("<<ComboboxSelected>>", self._on_type_changed)

        # 维护后状态
        tk.Label(r0, text="状态", fg=FG_PRIMARY, bg=BG_PANEL,
                 font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(12, 0))
        self.var_status = tk.StringVar(value="正常")
        self.cmb_status = ttk.Combobox(r0, textvariable=self.var_status,
                                  values=["正常", "需关注", "异常"], width=8,
                                  state="readonly")
        self.cmb_status.pack(side=tk.LEFT, padx=4)

        # 建议下次维护日期
        tk.Label(r0, text="下次维护", fg=FG_PRIMARY, bg=BG_PANEL,
                 font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(12, 0))
        self.var_next_date = tk.StringVar()
        self.ent_next_date = ttk.Entry(r0, textvariable=self.var_next_date, width=14)
        self.ent_next_date.pack(side=tk.LEFT, padx=4)

        ttk.Button(r0, text="自动计算", command=self._auto_next_date).pack(side=tk.LEFT, padx=(4, 0))

        # 备注行
        r1 = tk.Frame(panel, bg=BG_PANEL)
        r1.pack(fill=tk.X, pady=(4, 0))
        tk.Label(r1, text="备注", fg=FG_PRIMARY, bg=BG_PANEL,
                 font=("Segoe UI", 9)).pack(side=tk.LEFT)
        self.var_notes = tk.StringVar()
        self.ent_notes = ttk.Entry(r1, textvariable=self.var_notes, width=60)
        self.ent_notes.pack(side=tk.LEFT, padx=4, fill=tk.X, expand=True)

        # ---- 操作按钮行 ----
        btn_row = tk.Frame(panel, bg=BG_PANEL)
        btn_row.pack(fill=tk.X, pady=(8, 2))
        tk.Label(btn_row, text="", bg=BG_PANEL).pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Button(btn_row, text="清空", command=self._clear_form).pack(side=tk.RIGHT, padx=4)
        self.btn_cancel = ttk.Button(btn_row, text="取消编辑", command=self._cancel_edit, state=tk.DISABLED)
        self.btn_cancel.pack(side=tk.RIGHT, padx=4)
        ttk.Button(btn_row, text="保存维护记录", command=self._save, style="Primary.TButton").pack(side=tk.RIGHT)

        # 上次维护提示
        self.lbl_last_info = tk.Label(panel, text="", fg=FG_SECONDARY, bg=BG_PANEL,
                                       font=("Segoe UI", 8))
        self.lbl_last_info.pack(fill=tk.X, pady=(6, 0))

        # ---- 工具栏 ----
        bar = tk.Frame(self, bg=BG_DARK, padx=16, pady=8)
        bar.pack(fill=tk.X)
        tk.Label(bar, text="搜索:", fg=FG_PRIMARY, bg=BG_DARK,
                 font=("Segoe UI", 9)).pack(side=tk.LEFT)
        self.var_search = tk.StringVar()
        self.var_search.trace_add("write", lambda *a: self._apply_filter())
        ttk.Entry(bar, textvariable=self.var_search, width=30).pack(side=tk.LEFT, padx=4)

        ttk.Button(bar, text="导出 CSV", command=self._export_csv).pack(side=tk.RIGHT, padx=4)
        ttk.Button(bar, text="导出 Excel", command=self._export_excel).pack(side=tk.RIGHT, padx=4)
        ttk.Button(bar, text="删除选中", command=self._delete, style="Danger.TButton").pack(side=tk.RIGHT, padx=4)
        ttk.Button(bar, text="编辑选中", command=self._edit_selected).pack(side=tk.RIGHT, padx=4)

        # ---- 模型图例 ----
        legend = tk.Frame(self, bg=BG_DARK, padx=16)
        legend.pack(fill=tk.X)
        for text, color in [("● 正常", GREEN_STATUS), ("● 需关注", YELLOW_STATUS), ("● 异常", RED_STATUS)]:
            tk.Label(legend, text=text, fg=color, bg=BG_DARK,
                     font=("Segoe UI", 8)).pack(side=tk.LEFT, padx=(0, 12))

        # ---- 表格 ----
        tree_frame = tk.Frame(self, bg=BG_DARK, padx=16, pady=4)
        tree_frame.pack(fill=tk.BOTH, expand=True)

        cols = ("id", "日期", "姓名", "维护类型", "状态", "下次维护", "备注", "创建时间")
        self.tree = ttk.Treeview(tree_frame, columns=cols, show="headings",
                                 selectmode="extended")
        widths = [35, 95, 65, 170, 65, 100, 200, 150]
        for c, w in zip(cols, widths):
            self.tree.heading(c, text=c, command=lambda col=c: self._sort(col))
            self.tree.column(c, width=w, minwidth=30, stretch=c in ("备注", "维护类型"))

        vsb = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)

        self.tree.bind("<Double-1>", lambda e: self._edit_selected())

        # 右键拖拽多选
        self.tree.bind("<ButtonPress-3>", self._on_rclick_press)
        self.tree.bind("<B3-Motion>", self._on_rclick_drag)
        self.tree.bind("<ButtonRelease-3>", self._on_rclick_release)

        # 状态栏
        self.status = tk.Label(self, text="就绪", fg=FG_SECONDARY, bg=BG_DARK,
                                anchor=tk.W, padx=16, pady=4,
                                font=("Segoe UI", 8))
        self.status.pack(side=tk.BOTTOM, fill=tk.X)

        # Enter 键在输入字段间跳转
        self._bind_enter_nav()

    def _bind_enter_nav(self):
        """Enter 键在日期 → 姓名 → 维护类型 → 状态 → 下次维护 → 备注 之间跳转"""
        order = [
            self.ent_date, self.ent_name, self.cmb_type,
            self.cmb_status, self.ent_next_date, self.ent_notes,
        ]

        def make_handler(nxt):
            def handler(event):
                nxt.focus_set()
                if isinstance(nxt, ttk.Combobox):
                    try:
                        nxt.event_generate("<Button-1>")
                    except Exception:
                        pass
                else:
                    try:
                        nxt.select_range(0, tk.END)
                    except Exception:
                        pass
                return "break"
            return handler

        for i, ent in enumerate(order):
            nxt = order[(i + 1) % len(order)]
            ent.bind("<Return>", make_handler(nxt))

    def _on_type_changed(self, event=None):
        """选择维护类型后自动计算建议日期"""
        rt = self.var_record_type.get()
        if rt in self.mt_config:
            self._auto_next_date()

    def _auto_next_date(self):
        rt = self.var_record_type.get()
        if rt in self.mt_config:
            days = self.mt_config[rt]
            nd = datetime.now() + timedelta(days=days)
            self.var_next_date.set(nd.strftime("%Y-%m-%d"))

    def _load_data(self):
        with get_db() as conn:
            self._all_rows = conn.execute(
                "SELECT * FROM maintenance WHERE ms_type=? ORDER BY date DESC",
                (self.ms_type,)
            ).fetchall()
        self._apply_filter()

    def _apply_filter(self):
        q = self.var_search.get().lower()
        rows = self._all_rows if not q else [
            r for r in self._all_rows
            if any(q in str(v).lower() for v in dict(r).values())
        ]
        self.tree.delete(*self.tree.get_children())
        for r in rows:
            d = dict(r)
            status_val = d.get("status", "")
            iid = str(d["id"])
            self.tree.insert("", tk.END, iid=iid,
                             values=(d["id"], d["date"], d.get("name", ""),
                                     d["record_type"],
                                     status_val,
                                     d.get("next_date", ""),
                                     d.get("notes", ""),
                                     d.get("created_at", "")))

            # 状态颜色标记
            tag = f"tag_{iid}"
            if status_val == "正常":
                self.tree.tag_configure(tag, foreground=GREEN_STATUS)
            elif status_val == "需关注":
                self.tree.tag_configure(tag, foreground=YELLOW_STATUS)
            elif status_val == "异常":
                self.tree.tag_configure(tag, foreground=RED_STATUS)
            else:
                self.tree.tag_configure(tag, foreground=FG_PRIMARY)
            self.tree.item(iid, tags=(tag,))

            # 检查是否超期
            nd_str = d.get("next_date", "")
            if nd_str:
                try:
                    nd = datetime.strptime(nd_str, "%Y-%m-%d")
                    if nd < datetime.now():
                        self.tree.tag_configure(f"overdue_{iid}", background="#3d1518")
                        self.tree.item(iid, tags=(f"overdue_{iid}", tag))
                except ValueError:
                    pass

        self._set_status(f"共 {len(rows)} 条维护记录")
        self._update_last_info()

    def _update_last_info(self):
        """更新上次维护提示"""
        info_parts = []
        for mt_name, mt_days in MAINTENANCE_TYPES.get(self.ms_type, []):
            last = None
            for r in self._all_rows:
                if r["record_type"] == mt_name:
                    last = r["date"]
                    break
            if last:
                try:
                    ld = datetime.strptime(last, "%Y-%m-%d")
                    ago = (datetime.now() - ld).days
                    warn = " ⚠ 已超期" if ago > mt_days else ""
                    cycle = f"{mt_days}天" if mt_days > 0 else "定期"
                    info_parts.append(f"{mt_name}: 上次 {last}（{ago}天前 / 周期{cycle}）{warn}")
                except ValueError:
                    info_parts.append(f"{mt_name}: 上次 {last}")
            else:
                info_parts.append(f"{mt_name}: 暂无记录")
        self.lbl_last_info.config(text="  |  ".join(info_parts))

    def _save(self):
        data = {
            "date": self.var_date.get().strip(),
            "name": self.var_name.get().strip(),
            "record_type": self.var_record_type.get().strip(),
            "status": self.var_status.get().strip() or "正常",
            "notes": self.var_notes.get().strip(),
            "next_date": self.var_next_date.get().strip(),
        }
        if not data["date"]:
            messagebox.showwarning("提示", "请填写维护日期。", parent=self)
            return
        if not data["record_type"]:
            messagebox.showwarning("提示", "请选择维护类型。", parent=self)
            return

        try:
            with get_db() as conn:
                if self.editing_id:
                    conn.execute(
                        "UPDATE maintenance SET date=?, name=?, record_type=?, status=?, notes=?, next_date=? WHERE id=?",
                        (data["date"], data["name"], data["record_type"], data["status"],
                         data["notes"], data["next_date"], self.editing_id),
                    )
                    self._set_status(f"已更新维护记录 #{self.editing_id}")
                else:
                    cur = conn.execute(
                        "INSERT INTO maintenance (ms_type, date, name, record_type, status, notes, next_date) "
                        "VALUES (?, ?, ?, ?, ?, ?, ?)",
                        (self.ms_type, data["date"], data["name"], data["record_type"], data["status"],
                         data["notes"], data["next_date"]),
                    )
                    self._set_status(f"已创建维护记录 #{cur.lastrowid}")

            self._cancel_edit()
            self._clear_form()
            self._load_data()
        except Exception as e:
            messagebox.showerror("保存失败", f"数据库错误: {e}", parent=self)

    def _edit_selected(self):
        sel = self.tree.selection()
        if not sel:
            return
        rid = int(sel[0])
        row = next((r for r in self._all_rows if r["id"] == rid), None)
        if not row:
            return

        self.editing_id = rid
        self.var_date.set(row["date"])
        self.var_name.set(row["name"] if "name" in row.keys() else "")
        self.var_record_type.set(row["record_type"])
        self.var_status.set(row["status"] if "status" in row.keys() else "")
        self.var_notes.set(row["notes"] if "notes" in row.keys() else "")
        self.var_next_date.set(row["next_date"] if "next_date" in row.keys() else "")
        self.btn_cancel.config(state=tk.NORMAL)
        self._set_status(f"正在编辑维护记录 #{rid}")

    def _cancel_edit(self):
        self.editing_id = None
        self.btn_cancel.config(state=tk.DISABLED)
        self._set_status("就绪")

    def _delete(self):
        sel = self.tree.selection()
        if not sel:
            return
        count = len(sel)
        if not messagebox.askyesno("确认", f"确定要删除选中的 {count} 条维护记录吗？此操作不可撤销。"):
            return
        ids = [int(s) for s in sel]
        with get_db() as conn:
            conn.executemany("DELETE FROM maintenance WHERE id = ?", [(i,) for i in ids])
        if self.editing_id in ids:
            self._cancel_edit()
            self._clear_form()
        self._load_data()
        self._set_status(f"已删除 {count} 条维护记录")

    def _clear_form(self):
        self.var_date.set(datetime.now().strftime("%Y-%m-%d"))
        self.var_name.set("")
        self.var_record_type.set("")
        self.var_status.set("正常")
        self.var_notes.set("")
        self.var_next_date.set("")
        self.lbl_last_info.config(text="")

    def _sort(self, col):
        rev = self._sort_dir.get(col, False)
        self._sort_dir[col] = not rev
        key_map = {
            "id": lambda r: r["id"],
            "日期": lambda r: r["date"],
            "姓名": lambda r: r.get("name", ""),
            "维护类型": lambda r: r["record_type"],
            "状态": lambda r: r.get("status", ""),
            "下次维护": lambda r: r.get("next_date", ""),
            "备注": lambda r: r.get("notes", ""),
            "创建时间": lambda r: r.get("created_at", ""),
        }
        self._all_rows = sorted(self._all_rows, key=key_map.get(col, lambda r: r["id"]), reverse=rev)
        self._apply_filter()

    def _set_status(self, msg):
        self.status.config(text=msg)

    # ========== 右键拖拽多选 ==========
    def _on_rclick_press(self, event):
        item = self.tree.identify_row(event.y)
        if item:
            self._rclick_start = item
            self.tree.selection_set(item)
        return "break"

    def _on_rclick_drag(self, event):
        if self._rclick_start is None:
            return "break"
        item = self.tree.identify_row(event.y)
        if not item:
            return "break"
        all_items = self.tree.get_children()
        try:
            start_idx = all_items.index(self._rclick_start)
            cur_idx = all_items.index(item)
        except ValueError:
            return "break"
        lo, hi = min(start_idx, cur_idx), max(start_idx, cur_idx)
        self.tree.selection_set(*all_items[lo:hi + 1])
        return "break"

    def _on_rclick_release(self, event):
        self._rclick_start = None

    # ========== 导出 ==========
    def _export_csv(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("提示", "请先在表格中选中要导出的记录（可使用鼠标拖拽多选）。", parent=self)
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV 文件", "*.csv")],
            initialfile=f"maintenance_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            parent=self,
        )
        if not path:
            return

        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["ID", "日期", "姓名", "维护类型", "状态", "下次维护", "备注", "创建时间", "质谱类型"])
            for iid in sel:
                values = list(self.tree.item(iid, "values"))
                rid = int(iid)
                row = next((r for r in self._all_rows if r["id"] == rid), None)
                created_at = row["created_at"] if row else ""
                ms_type = row["ms_type"] if row else self.ms_type
                values.append(created_at)
                values.append(ms_type)
                w.writerow(values)
        self._set_status(f"已导出 {len(sel)} 条维护记录到 {os.path.basename(path)}")

    def _export_excel(self):
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showwarning("提示", "需要安装 openpyxl 库才能导出 Excel。\n请在命令行运行: pip install openpyxl", parent=self)
            return

        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("提示", "请先在表格中选中要导出的记录（可使用鼠标拖拽多选）。", parent=self)
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")],
            initialfile=f"maintenance_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            parent=self,
        )
        if not path:
            return

        ids = [int(iid) for iid in sel]
        created_at_map = {}
        try:
            with get_db() as conn:
                placeholders = ",".join("?" * len(ids))
                rows = conn.execute(
                    f"SELECT id, created_at FROM maintenance WHERE id IN ({placeholders})",
                    ids,
                ).fetchall()
                created_at_map = {r["id"]: r["created_at"] for r in rows}
        except Exception:
            pass

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{self.ms_type} 维护记录"

        headers = ["ID", "日期", "姓名", "维护类型", "状态", "下次维护", "备注", "创建时间", "质谱类型"]
        ws.append(headers)

        for iid in sel:
            values = list(self.tree.item(iid, "values"))
            rid = int(iid)
            values.append(created_at_map.get(rid, ""))
            values.append(self.ms_type)
            ws.append(values)
            # 状态列(第5列)按状态着色：正常=绿色, 需关注=黄色, 异常=红色
            row_num = ws.max_row
            status_val = values[4]
            color_map = {"正常": "3FB950", "需关注": "D2991D", "异常": "F85149"}
            if status_val in color_map:
                ws.cell(row=row_num, column=5).font = openpyxl.styles.Font(bold=True, color=color_map[status_val])

        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_len = 0
            for cell in col_cells:
                if cell.value is not None:
                    cell_str = str(cell.value)
                    char_len = sum(2 if ord(ch) > 127 else 1 for ch in cell_str)
                    max_len = max(max_len, char_len)
            adjusted = min(max_len + 4, 80)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted

        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)

        wb.save(path)
        self._set_status(f"已导出 {len(sel)} 条维护记录到 {os.path.basename(path)}")


# ============================================================
# 主界面 App
# ============================================================
class App:
    def __init__(self, ms_type="Q-IM-TOF"):
        self.ms_type = ms_type
        self.ms_cfg = MS_CONFIG.get(ms_type, MS_CONFIG["Q-IM-TOF"])
        self.editing_id = None
        self._sort_dir = {}
        self._all_rows = []
        self._rclick_start = None

        init_db()
        self.config = load_config()

        self.root = tk.Tk()
        self.root.title(f"MS Recording & Maintenance — {ms_type}")
        self.root.geometry("1350x880")
        self.root.minsize(1000, 650)
        self.root.configure(bg=BG_DARK)

        self.style = apply_dark_theme(self.root)
        self._build_ui()
        self._load_data()
        self._bind_enter_nav()

    def _bind_enter_nav(self):
        self._input_order = [
            self.ent_date, self.ent_timeperiod, self.ent_name, self.ent_purpose,
            self.ent_solvent, self.cmb_cleaned,
            self.ent_ionsrc, self.ent_sinfo, self.ent_charge, self.ent_speaks,
        ]

        def make_handler(nxt):
            def handler(event):
                nxt.focus_set()
                if isinstance(nxt, ttk.Combobox):
                    try:
                        nxt.event_generate("<Button-1>")
                    except Exception:
                        pass
                elif nxt is self.ent_ionsrc:
                    try:
                        nxt.icursor(tk.END)
                    except Exception:
                        pass
                else:
                    try:
                        nxt.select_range(0, tk.END)
                    except Exception:
                        pass
                return "break"
            return handler

        for i, ent in enumerate(self._input_order):
            nxt = self._input_order[(i + 1) % len(self._input_order)]
            ent.bind("<Return>", make_handler(nxt))

    def _build_ui(self):
        # ---- 顶部装饰条 ----
        canvas = tk.Canvas(self.root, height=3, bg=BG_DARK, highlightthickness=0)
        canvas.pack(fill=tk.X)
        canvas.create_rectangle(0, 0, 1400, 3, fill=ACCENT, outline="")
        canvas.create_rectangle(0, 0, 800, 3, fill=ACCENT_HIGHLIGHT, outline="")

        # ---- 顶部工具栏 ----
        toolbar = tk.Frame(self.root, bg=BG_DARK, padx=20, pady=14)
        toolbar.pack(fill=tk.X)

        # 标题区
        title_frame = tk.Frame(toolbar, bg=BG_DARK)
        title_frame.pack(side=tk.LEFT)

        tk.Label(title_frame, text="MS Recording & Maintenance",
                 font=("Segoe UI", 16, "bold"), fg=ACCENT_HIGHLIGHT, bg=BG_DARK).pack(anchor=tk.W)
        tk.Label(title_frame, text=f"当前仪器: {self.ms_type}",
                 font=("Segoe UI", 9), fg=FG_SECONDARY, bg=BG_DARK).pack(anchor=tk.W)

        # 状态指示灯
        status_frame = tk.Frame(toolbar, bg=BG_DARK)
        status_frame.pack(side=tk.LEFT, padx=30)

        self.lbl_status_dot = tk.Label(status_frame, text="●", fg=GREEN_STATUS,
                                        bg=BG_DARK, font=("Segoe UI", 10))
        self.lbl_status_dot.pack(side=tk.LEFT)
        tk.Label(status_frame, text="系统就绪", fg=FG_SECONDARY, bg=BG_DARK,
                 font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(4, 0))

        # 按钮区
        btn_frame = tk.Frame(toolbar, bg=BG_DARK)
        btn_frame.pack(side=tk.RIGHT)

        btn_switch = tk.Button(btn_frame, text="⇄ 切换质谱", width=12,
                               font=("Segoe UI", 9, "bold"),
                               bg=BG_INPUT, fg=ACCENT,
                               activebackground="#30363d", activeforeground=ACCENT_HIGHLIGHT,
                               relief="flat", bd=1, cursor="hand2",
                               command=self._switch_ms_type)
        btn_switch.pack(side=tk.RIGHT, padx=4)
        btn_switch.bind("<Enter>", lambda e: btn_switch.configure(bg="#1c2638"))
        btn_switch.bind("<Leave>", lambda e: btn_switch.configure(bg=BG_INPUT))

        btn_maint = tk.Button(btn_frame, text="🔧 维护记录", width=12,
                              font=("Segoe UI", 9, "bold"),
                              bg=ACCENT, fg="#ffffff",
                              activebackground=ACCENT_HIGHLIGHT, activeforeground="#ffffff",
                              relief="flat", bd=1, cursor="hand2",
                              command=self._open_maintenance)
        btn_maint.pack(side=tk.RIGHT, padx=4)

        ttk.Separator(self.root, orient=tk.HORIZONTAL).pack(fill=tk.X)

        # ---- 表单区 ----
        form_panel = tk.Frame(self.root, bg=BG_PANEL)
        form_panel.pack(fill=tk.X, padx=16, pady=(12, 0))

        top = tk.Frame(form_panel, bg=BG_PANEL, padx=16, pady=14)
        top.pack(fill=tk.X)

        # Row 0: 日期 | 时间段 | 姓名
        r0 = tk.Frame(top, bg=BG_PANEL)
        r0.pack(fill=tk.X, pady=(0, 8))

        tk.Label(r0, text="日期", fg=FG_PRIMARY, bg=BG_PANEL,
                 font=("Segoe UI", 9)).pack(side=tk.LEFT)
        self.var_date = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        self.ent_date = ttk.Entry(r0, textvariable=self.var_date, width=14)
        self.ent_date.pack(side=tk.LEFT, padx=(4, 18))

        tk.Label(r0, text="时间段", fg=FG_PRIMARY, bg=BG_PANEL,
                 font=("Segoe UI", 9)).pack(side=tk.LEFT)
        self.var_timeperiod = tk.StringVar()
        self.ent_timeperiod = ttk.Entry(r0, textvariable=self.var_timeperiod, width=16)
        self.ent_timeperiod.pack(side=tk.LEFT, padx=4)
        tk.Label(r0, text="例: 6:45-9:00", fg=FG_SECONDARY, bg=BG_PANEL,
                 font=("Segoe UI", 8)).pack(side=tk.LEFT, padx=(0, 20))

        tk.Label(r0, text="姓名", fg=FG_PRIMARY, bg=BG_PANEL,
                 font=("Segoe UI", 9)).pack(side=tk.LEFT)
        self.var_name = tk.StringVar()
        self.ent_name = ttk.Entry(r0, textvariable=self.var_name, width=16)
        self.ent_name.pack(side=tk.LEFT, padx=4)

        ttk.Button(r0, text="清空表单", command=self._clear_form).pack(side=tk.RIGHT, padx=4)
        ttk.Button(r0, text="保存", command=self._save, style="Primary.TButton").pack(side=tk.RIGHT)
        self.btn_cancel = ttk.Button(r0, text="取消编辑", command=self._cancel_edit, state=tk.DISABLED)
        self.btn_cancel.pack(side=tk.RIGHT, padx=4)

        # Row 1: 实验目的
        r1 = tk.Frame(top, bg=BG_PANEL)
        r1.pack(fill=tk.X, pady=(0, 8))
        tk.Label(r1, text="实验目的", fg=FG_PRIMARY, bg=BG_PANEL,
                 font=("Segoe UI", 9)).pack(side=tk.LEFT)
        self.var_purpose = tk.StringVar()
        self.ent_purpose = ttk.Entry(r1, textvariable=self.var_purpose, width=100)
        self.ent_purpose.pack(side=tk.LEFT, padx=4, fill=tk.X, expand=True)

        # Row 1.5: 溶剂 | 是否清洗干净
        r1b = tk.Frame(top, bg=BG_PANEL)
        r1b.pack(fill=tk.X, pady=(0, 8))
        tk.Label(r1b, text="溶剂", fg=FG_PRIMARY, bg=BG_PANEL,
                 font=("Segoe UI", 9)).pack(side=tk.LEFT)
        self.var_solvent = tk.StringVar()
        self.ent_solvent = ttk.Entry(r1b, textvariable=self.var_solvent, width=30)
        self.ent_solvent.pack(side=tk.LEFT, padx=4)

        tk.Label(r1b, text="是否清洗干净", fg=FG_PRIMARY, bg=BG_PANEL,
                 font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(20, 0))
        self.var_cleaned = tk.StringVar()
        self.cmb_cleaned = ttk.Combobox(r1b, textvariable=self.var_cleaned,
                                         values=["是", "否"], width=6, state="readonly")
        self.cmb_cleaned.pack(side=tk.LEFT, padx=4)

        # Row 2: 测试条件
        r2 = tk.Frame(top, bg=BG_PANEL)
        r2.pack(fill=tk.X, pady=(0, 8))
        tk.Label(r2, text=self.ms_cfg["ion_label"], fg=FG_PRIMARY, bg=BG_PANEL,
                 font=("Segoe UI", 9)).pack(side=tk.LEFT)
        self.var_ionsrc = tk.StringVar(value=self.ms_cfg["ion_default"])
        self.ent_ionsrc = ttk.Entry(r2, textvariable=self.var_ionsrc, width=70)
        self.ent_ionsrc.pack(side=tk.LEFT, padx=4, fill=tk.X, expand=True)

        # Row 3: 样品信息 | 电荷 | [计算]
        r3 = tk.Frame(top, bg=BG_PANEL)
        r3.pack(fill=tk.X, pady=(0, 8))
        tk.Label(r3, text="样品信息", fg=FG_PRIMARY, bg=BG_PANEL,
                 font=("Segoe UI", 9)).pack(side=tk.LEFT)
        self.var_sinfo = tk.StringVar()
        self.ent_sinfo = ttk.Entry(r3, textvariable=self.var_sinfo, width=55)
        self.ent_sinfo.pack(side=tk.LEFT, padx=4, fill=tk.X, expand=True)

        tk.Label(r3, text="电荷", fg=FG_PRIMARY, bg=BG_PANEL,
                 font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(14, 0))
        self.var_charge = tk.StringVar()
        self.ent_charge = ttk.Entry(r3, textvariable=self.var_charge, width=8)
        self.ent_charge.pack(side=tk.LEFT, padx=4)
        tk.Label(r3, text="例: 2+, 1-, +3", fg=FG_SECONDARY, bg=BG_PANEL,
                 font=("Segoe UI", 8)).pack(side=tk.LEFT, padx=(0, 8))

        ttk.Button(r3, text="计算样品峰 m/z", command=self._calc_peaks).pack(side=tk.LEFT)

        # Row 4: 样品峰
        r4 = tk.Frame(top, bg=BG_PANEL)
        r4.pack(fill=tk.X)
        tk.Label(r4, text="样品峰", fg=FG_PRIMARY, bg=BG_PANEL,
                 font=("Segoe UI", 9)).pack(side=tk.LEFT)
        self.var_speaks = tk.StringVar()
        self.ent_speaks = ttk.Entry(r4, textvariable=self.var_speaks, width=100)
        self.ent_speaks.pack(side=tk.LEFT, padx=4, fill=tk.X, expand=True)

        ttk.Separator(self.root, orient=tk.HORIZONTAL).pack(fill=tk.X)

        # ---- 表格工具栏 ----
        bar = tk.Frame(self.root, bg=BG_DARK, padx=16, pady=10)
        bar.pack(fill=tk.X)
        tk.Label(bar, text="搜索:", fg=FG_PRIMARY, bg=BG_DARK,
                 font=("Segoe UI", 9)).pack(side=tk.LEFT)
        self.var_search = tk.StringVar()
        self.var_search.trace_add("write", lambda *a: self._apply_filter())
        ttk.Entry(bar, textvariable=self.var_search, width=30).pack(side=tk.LEFT, padx=4)

        ttk.Button(bar, text="导出 CSV", command=self._export_csv).pack(side=tk.RIGHT, padx=4)
        ttk.Button(bar, text="导出 Excel", command=self._export_excel).pack(side=tk.RIGHT, padx=4)
        ttk.Button(bar, text="删除选中", command=self._delete, style="Danger.TButton").pack(side=tk.RIGHT, padx=4)
        ttk.Button(bar, text="编辑选中", command=self._edit_selected).pack(side=tk.RIGHT, padx=4)

        # ---- 表格 ----
        tree_frame = tk.Frame(self.root, bg=BG_DARK, padx=16, pady=4)
        tree_frame.pack(fill=tk.BOTH, expand=True)

        cols = ("id", "日期", "时间段", "实验目的", "姓名", "测试条件", "样品信息", "电荷", "样品峰", "溶剂", "是否清洗干净")
        self.tree = ttk.Treeview(tree_frame, columns=cols, show="headings", selectmode="extended")
        widths = [35, 95, 85, 160, 65, 160, 130, 50, 130, 70, 80]
        for c, w in zip(cols, widths):
            self.tree.heading(c, text=c, command=lambda col=c: self._sort(col))
            self.tree.column(c, width=w, minwidth=35, stretch=c not in ("id", "电荷"))

        vsb = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)

        self.tree.bind("<Double-1>", lambda e: self._edit_selected())
        self.root.bind("<Control-s>", lambda e: self._save())
        self.root.bind("<Control-f>", lambda e: self._focus_search())

        # 右键拖拽多选
        self.tree.bind("<ButtonPress-3>", self._on_rclick_press)
        self.tree.bind("<B3-Motion>", self._on_rclick_drag)
        self.tree.bind("<ButtonRelease-3>", self._on_rclick_release)

        # 状态栏
        self.status = tk.Label(self.root, text="就绪", fg=FG_SECONDARY, bg=BG_DARK,
                                anchor=tk.W, padx=16, pady=4,
                                font=("Segoe UI", 8))
        self.status.pack(side=tk.BOTTOM, fill=tk.X)

    # ========== 表单操作 ==========
    def _focus_search(self):
        for child in self.root.winfo_children():
            if isinstance(child, tk.Frame):
                for w in child.winfo_children():
                    if isinstance(w, ttk.Entry):
                        try:
                            if w.cget("textvariable") == str(self.var_search):
                                w.focus_set()
                                return
                        except Exception:
                            pass

    def _on_rclick_press(self, event):
        item = self.tree.identify_row(event.y)
        if item:
            self._rclick_start = item
            self.tree.selection_set(item)
        return "break"

    def _on_rclick_drag(self, event):
        if self._rclick_start is None:
            return "break"
        item = self.tree.identify_row(event.y)
        if not item:
            return "break"
        all_items = self.tree.get_children()
        try:
            start_idx = all_items.index(self._rclick_start)
            cur_idx = all_items.index(item)
        except ValueError:
            return "break"
        lo, hi = min(start_idx, cur_idx), max(start_idx, cur_idx)
        self.tree.selection_set(*all_items[lo:hi + 1])
        return "break"

    def _on_rclick_release(self, event):
        self._rclick_start = None

    # ========== m/z 计算 ==========
    def _calc_peaks(self):
        sinfo = self.var_sinfo.get().strip()
        charge = self.var_charge.get().strip()
        if not sinfo:
            messagebox.showwarning("提示", "请先输入样品信息（分子式）。")
            return
        if not charge:
            messagebox.showwarning("提示", "请先输入电荷信息（如 2+, 1-）。")
            return
        try:
            mz, err = calc_mz(sinfo, charge)
        except ValueError as e:
            messagebox.showerror("分子式解析错误", str(e))
            return
        if err:
            messagebox.showerror("错误", err)
            return
        self.var_speaks.set(f"m/z {mz:.2f}")
        self._set_status(f"已计算: 质量={parse_formula(sinfo)[0]:.4f}, 电荷={charge} → m/z={mz:.2f}")

    # ========== CRUD ==========
    def _save(self):
        data = {
            "date": self.var_date.get().strip(),
            "time_period": self.var_timeperiod.get().strip(),
            "purpose": self.var_purpose.get().strip(),
            "name": self.var_name.get().strip(),
            "ion_source": self.var_ionsrc.get().strip(),
            "sample_info": self.var_sinfo.get().strip(),
            "charge": self.var_charge.get().strip(),
            "sample_peaks": self.var_speaks.get().strip(),
            "solvent": self.var_solvent.get().strip(),
            "cleaned": self.var_cleaned.get().strip(),
        }
        if not all(data.values()):
            messagebox.showwarning("提示", "所有字段均为必填，请完整填写。")
            return

        with get_db() as conn:
            if self.editing_id:
                conn.execute(
                    "UPDATE experiments SET date=?, time_period=?, purpose=?, name=?, "
                    "ion_source=?, sample_info=?, charge=?, sample_peaks=?, solvent=?, cleaned=?, ms_type=? WHERE id=?",
                    (*data.values(), self.ms_type, self.editing_id),
                )
                self._set_status(f"已更新记录 #{self.editing_id}")
            else:
                cur = conn.execute(
                    "INSERT INTO experiments (date, time_period, purpose, name, ion_source, sample_info, charge, sample_peaks, solvent, cleaned, ms_type) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (*data.values(), self.ms_type),
                )
                self._set_status(f"已创建记录 #{cur.lastrowid}")

        self._cancel_edit()
        self._clear_form()
        self._load_data()

    def _load_data(self):
        with get_db() as conn:
            self._all_rows = conn.execute(
                "SELECT * FROM experiments WHERE ms_type=? ORDER BY date DESC, time_period DESC",
                (self.ms_type,)
            ).fetchall()
        self._apply_filter()

    def _apply_filter(self):
        q = self.var_search.get().lower()
        rows = self._all_rows if not q else [
            r for r in self._all_rows
            if any(q in str(v).lower() for v in dict(r).values())
        ]
        self.tree.delete(*self.tree.get_children())
        for r in rows:
            d = dict(r)
            self.tree.insert("", tk.END, iid=str(d["id"]),
                             values=(d["id"], d["date"], d["time_period"], d["purpose"],
                                     d["name"], d["ion_source"], d["sample_info"],
                                     d["charge"], d["sample_peaks"],
                                     d.get("solvent", ""), d.get("cleaned", "")))
        self._auto_fit_columns()
        self._set_status(f"共 {len(rows)} 条记录")

    def _auto_fit_columns(self):
        import tkinter.font as tkfont
        try:
            style = ttk.Style()
            font_name = style.lookup(self.tree["style"], "font")
            font = tkfont.Font(font=font_name) if font_name else tkfont.nametofont("TkDefaultFont")
        except Exception:
            font = tkfont.nametofont("TkDefaultFont")
        padding = 28

        cols = ("id", "日期", "时间段", "实验目的", "姓名", "测试条件", "样品信息", "电荷", "样品峰", "溶剂", "是否清洗干净")
        children = self.tree.get_children()

        for col in cols:
            max_width = font.measure(col) + padding
            col_idx = cols.index(col)
            for child in children:
                values = self.tree.item(child, "values")
                if col_idx < len(values):
                    cell_text = str(values[col_idx])
                    max_width = max(max_width, font.measure(cell_text) + padding)
            max_width = min(max_width, 600)
            min_w = self.tree.column(col)["minwidth"]
            self.tree.column(col, width=max(max_width, min_w))

    def _edit_selected(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("提示", "请先在表格中选中一条记录。")
            return
        rid = int(sel[0])
        row = next((r for r in self._all_rows if r["id"] == rid), None)
        if not row:
            return

        self.editing_id = rid
        self.var_date.set(row["date"])
        self.var_timeperiod.set(row["time_period"])
        self.var_purpose.set(row["purpose"])
        self.var_name.set(row["name"])
        self.var_ionsrc.set(row["ion_source"])
        self.var_sinfo.set(row["sample_info"])
        self.var_charge.set(row["charge"])
        self.var_speaks.set(row["sample_peaks"])
        self.var_solvent.set(row["solvent"] if "solvent" in row.keys() else "")
        self.var_cleaned.set(row["cleaned"] if "cleaned" in row.keys() else "")
        self.btn_cancel.config(state=tk.NORMAL)
        self._set_status(f"正在编辑记录 #{rid}")

    def _cancel_edit(self):
        self.editing_id = None
        self.btn_cancel.config(state=tk.DISABLED)
        self._set_status("就绪")

    def _delete(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("提示", "请先在表格中选中要删除的记录。")
            return
        count = len(sel)
        if not messagebox.askyesno("确认", f"确定要删除选中的 {count} 条记录吗？此操作不可撤销。"):
            return
        ids = [int(s) for s in sel]
        with get_db() as conn:
            conn.executemany("DELETE FROM experiments WHERE id = ?", [(i,) for i in ids])
        if self.editing_id in ids:
            self._cancel_edit()
            self._clear_form()
        self._load_data()
        self._set_status(f"已删除 {count} 条记录")

    def _clear_form(self):
        self.var_date.set(datetime.now().strftime("%Y-%m-%d"))
        self.var_timeperiod.set("")
        self.var_purpose.set("")
        self.var_name.set("")
        self.var_ionsrc.set(self.ms_cfg["ion_default"])
        self.var_sinfo.set("")
        self.var_charge.set("")
        self.var_speaks.set("")
        self.var_solvent.set("")
        self.var_cleaned.set("")

    # ========== 排序 ==========
    def _sort(self, col):
        rev = self._sort_dir.get(col, False)
        self._sort_dir[col] = not rev
        key_map = {
            "id": lambda r: r["id"],
            "日期": lambda r: r["date"],
            "时间段": lambda r: r["time_period"],
            "实验目的": lambda r: r["purpose"],
            "姓名": lambda r: r["name"],
            "测试条件": lambda r: r["ion_source"],
            "样品信息": lambda r: r["sample_info"],
            "电荷": lambda r: r["charge"],
            "样品峰": lambda r: r["sample_peaks"],
            "溶剂": lambda r: r.get("solvent", ""),
            "是否清洗干净": lambda r: r.get("cleaned", ""),
        }
        self._all_rows = sorted(self._all_rows, key=key_map.get(col, lambda r: r["id"]), reverse=rev)
        self._apply_filter()

    # ========== 导出 ==========
    def _export_csv(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("提示", "请先在表格中选中要导出的记录（可使用鼠标拖拽多选）。")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV 文件", "*.csv")],
            initialfile=f"experiments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
        )
        if not path:
            return

        ids = [int(iid) for iid in sel]
        created_at_map = {}
        try:
            with get_db() as conn:
                placeholders = ",".join("?" * len(ids))
                rows = conn.execute(
                    f"SELECT id, created_at FROM experiments WHERE id IN ({placeholders})",
                    ids,
                ).fetchall()
                created_at_map = {r["id"]: r["created_at"] for r in rows}
        except Exception:
            pass

        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["ID", "日期", "时间段", "实验目的", "姓名", "测试条件", "样品信息", "电荷", "样品峰", "溶剂", "是否清洗干净", "创建时间", "质谱类型"])
            for iid in sel:
                values = list(self.tree.item(iid, "values"))
                rid = int(iid)
                values.append(created_at_map.get(rid, ""))
                values.append(self.ms_type)
                w.writerow(values)
        self._set_status(f"已导出 {len(sel)} 条记录到 {os.path.basename(path)}")

    def _export_excel(self):
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showwarning("提示", "需要安装 openpyxl 库才能导出 Excel。\n请在命令行运行: pip install openpyxl")
            return

        sel = self.tree.selection()
        if not sel:
            sel = self.tree.get_children()
        if not sel:
            messagebox.showinfo("提示", "表格中没有数据可以导出。")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")],
            initialfile=f"experiments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{self.ms_type} Operation Log"

        headers = ["ID", "日期", "时间段", "实验目的", "姓名", "测试条件", "样品信息", "电荷", "样品峰", "溶剂", "是否清洗干净", "创建时间", "质谱类型"]
        ws.append(headers)

        ids = [int(iid) for iid in sel]
        created_at_map = {}
        try:
            with get_db() as conn:
                placeholders = ",".join("?" * len(ids))
                rows = conn.execute(
                    f"SELECT id, created_at FROM experiments WHERE id IN ({placeholders})",
                    ids,
                ).fetchall()
                created_at_map = {r["id"]: r["created_at"] for r in rows}
        except Exception:
            pass

        for iid in sel:
            values = list(self.tree.item(iid, "values"))
            rid = int(iid)
            values.append(created_at_map.get(rid, ""))
            values.append(self.ms_type)
            ws.append(values)

        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_len = 0
            for cell in col_cells:
                if cell.value is not None:
                    cell_str = str(cell.value)
                    char_len = sum(2 if ord(ch) > 127 else 1 for ch in cell_str)
                    max_len = max(max_len, char_len)
            adjusted = min(max_len + 4, 80)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted

        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)

        wb.save(path)
        self._set_status(f"已导出 {len(sel)} 条记录到 {os.path.basename(path)}")

    # ========== 切换质谱 / 维护记录 ==========
    def _switch_ms_type(self):
        if messagebox.askyesno("切换质谱", "切换质谱类型将重启应用，未保存的数据将丢失。\n确定继续吗？"):
            self.config.pop("ms_type", None)
            save_config(self.config)
            # 销毁当前窗口
            self.root.destroy()
            # 重新运行整个应用（进程内重新执行入口逻辑）
            run_app()

    def _open_maintenance(self):
        MaintenanceWindow(self.root, self.ms_type)

    def _set_status(self, msg):
        self.status.config(text=msg)

    def run(self):
        self.root.mainloop()


# ============================================================
# 入口
# ============================================================
def run_app():
    """运行应用的入口函数，支持进程内重启"""
    cfg = load_config()
    ms_type = cfg.get("ms_type")

    if not ms_type:
        root = tk.Tk()
        root.withdraw()
        sel = TypeSelector()
        root.wait_window(sel)
        root.destroy()
        ms_type = sel.result

    if ms_type:
        App(ms_type).run()


if __name__ == "__main__":
    run_app()

