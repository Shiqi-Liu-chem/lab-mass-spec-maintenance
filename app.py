import os
import re
import sys
import csv
import json
import sqlite3
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime

# ---- 数据库路径 ----
if getattr(sys, "frozen", False):
    APP_DIR = os.path.dirname(sys.executable)
else:
    APP_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(APP_DIR, "experiments.db")
CONFIG_PATH = os.path.join(APP_DIR, "config.json")

# ---- 原子量 ----
ATOMIC_MASSES = {
    "H": 1.00794, "He": 4.002602, "Li": 6.941, "Be": 9.012182, "B": 10.811,
    "C": 12.0107, "N": 14.0067, "O": 15.9994, "F": 18.9984032, "Ne": 20.1797,
    "Na": 22.98977, "Mg": 24.305, "Al": 26.98154, "Si": 28.0855, "P": 30.97376,
    "S": 32.065, "Cl": 35.453, "Ar": 39.948, "K": 39.0983, "Ca": 40.078,
    "Sc": 44.9559, "Ti": 47.867, "V": 50.9415, "Cr": 51.9961, "Mn": 54.938,
    "Fe": 55.845, "Co": 58.9332, "Ni": 58.6934, "Cu": 63.546, "Zn": 65.39,
    "Ga": 69.723, "Ge": 72.64, "As": 74.9216, "Se": 78.96, "Br": 79.904,
    "Kr": 83.8, "Rb": 85.4678, "Sr": 87.62, "Y": 88.9059, "Zr": 91.224,
    "Nb": 92.9064, "Mo": 95.94, "Tc": 98.0, "Ru": 101.07, "Rh": 102.9055,
    "Pd": 106.42, "Ag": 107.8682, "Cd": 112.411, "In": 114.818, "Sn": 118.71,
    "Sb": 121.76, "Te": 127.6, "I": 126.9045, "Xe": 131.293, "Cs": 132.9055,
    "Ba": 137.327, "La": 138.9055, "Ce": 140.116, "Pr": 140.9077, "Nd": 144.24,
    "Pm": 145.0, "Sm": 150.36, "Eu": 151.964, "Gd": 157.25, "Tb": 158.9253,
    "Dy": 162.5, "Ho": 164.9303, "Er": 167.259, "Tm": 168.9342, "Yb": 173.04,
    "Lu": 174.967, "Hf": 178.49, "Ta": 180.9479, "W": 183.84, "Re": 186.207,
    "Os": 190.23, "Ir": 192.217, "Pt": 195.078, "Au": 196.9665, "Hg": 200.59,
    "Tl": 204.3833, "Pb": 207.2, "Bi": 208.9804, "Po": 209.0, "At": 210.0,
    "Rn": 222.0, "Fr": 223.0, "Ra": 226.0, "Ac": 227.0, "Th": 232.0381,
    "Pa": 231.0359, "U": 238.0289, "Np": 237.0, "Pu": 244.0, "Am": 243.0,
    "Cm": 247.0, "Bk": 247.0, "Cf": 251.0, "Es": 252.0, "Fm": 257.0,
    "Md": 258.0, "No": 259.0, "Lr": 262.0, "Rf": 261.0, "Db": 262.0,
    "Sg": 266.0, "Bh": 264.0, "Hs": 277.0, "Mt": 268.0,
}

# ---- 分子式解析 & m/z ----
def parse_number(s, i):
    if i >= len(s) or not s[i].isdigit():
        return 1, i
    num = 0
    while i < len(s) and s[i].isdigit():
        num = num * 10 + int(s[i])
        i += 1
    return num, i


def parse_formula(formula, i=0):
    total = 0.0
    while i < len(formula):
        c = formula[i]
        if c in "([{":
            sub_mass, i = parse_formula(formula, i + 1)
            if i >= len(formula) or formula[i] not in ")]}":
                raise ValueError(f"括号不匹配，位置 {i}")
            i += 1
            num, i = parse_number(formula, i)
            total += sub_mass * num
        elif c in ")]}":
            return total, i
        elif c == "·" or c == ".":
            i += 1
        elif c.isupper():
            elem = c
            i += 1
            while i < len(formula) and formula[i].islower():
                elem += formula[i]
                i += 1
            num, i = parse_number(formula, i)
            mass = ATOMIC_MASSES.get(elem)
            if mass is None:
                raise ValueError(f"未知元素: {elem}")
            total += mass * num
        else:
            i += 1
    return total, i


def calc_mz(sample_info, charge_str):
    mass = parse_formula(sample_info)[0]
    charge_str = charge_str.strip()
    if not charge_str:
        return None, "请填写电荷信息"
    s = charge_str
    sign = 1
    if s.endswith("+") or s.endswith("-"):
        if s.endswith("-"):
            sign = -1
        num_str = s[:-1]
    elif s.startswith("+") or s.startswith("-"):
        if s.startswith("-"):
            sign = -1
        num_str = s[1:]
    else:
        return None, f"无法解析电荷: {charge_str}"
    z = 1 if num_str == "" else int(num_str)
    return mass / z, None


# ---- 配置读写 ----
def load_config():
    try:
        with open(CONFIG_PATH) as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def save_config(cfg):
    with open(CONFIG_PATH, "w") as f:
        json.dump(cfg, f, indent=2)


# ---- 数据库 ----
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn


def init_db():
    with get_db() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS experiments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT NOT NULL DEFAULT '',
                time_period TEXT NOT NULL DEFAULT '',
                purpose TEXT NOT NULL DEFAULT '',
                name TEXT NOT NULL DEFAULT '',
                ion_source TEXT NOT NULL DEFAULT '',
                sample_info TEXT NOT NULL DEFAULT '',
                charge TEXT NOT NULL DEFAULT '',
                sample_peaks TEXT NOT NULL DEFAULT '',
                solvent TEXT NOT NULL DEFAULT '',
                cleaned TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL DEFAULT (datetime('now', 'localtime'))
            )
        """)
        new_cols = [
            ("charge", "TEXT NOT NULL DEFAULT ''"),
            ("time_period", "TEXT NOT NULL DEFAULT ''"),
            ("solvent", "TEXT NOT NULL DEFAULT ''"),
            ("cleaned", "TEXT NOT NULL DEFAULT ''"),
        ]
        for col, dtype in new_cols:
            try:
                conn.execute(f"ALTER TABLE experiments ADD COLUMN {col} {dtype}")
            except sqlite3.OperationalError:
                pass


# ============================================================
#  ESI 屏幕捕获
# ============================================================
class RegionSelector(tk.Toplevel):
    """全屏半透明选区工具"""
    def __init__(self, master):
        super().__init__(master)
        self.attributes("-fullscreen", True, "-alpha", 0.35, "-topmost", True)
        self.configure(bg="black")
        self.overrideredirect(True)
        self.start_x = self.start_y = None
        self.rect = None
        self.result = None

        self.canvas = tk.Canvas(self, bg="black", highlightthickness=0)
        self.canvas.pack(fill=tk.BOTH, expand=True)
        self.canvas.bind("<ButtonPress-1>", self._on_press)
        self.canvas.bind("<B1-Motion>", self._on_drag)
        self.canvas.bind("<ButtonRelease-1>", self._on_release)
        self.bind("<Escape>", lambda e: self.destroy())

        # 提示文字
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        self.canvas.create_text(
            sw // 2, sh // 2, text="拖拽框选 ESI 数据区域\n按 Esc 取消",
            fill="white", font=("", 28), justify="center"
        )

    def _on_press(self, event):
        self.start_x, self.start_y = event.x, event.y
        if self.rect:
            self.canvas.delete(self.rect)

    def _on_drag(self, event):
        if self.rect:
            self.canvas.delete(self.rect)
        self.rect = self.canvas.create_rectangle(
            self.start_x, self.start_y, event.x, event.y,
            outline="#00ff00", width=2, dash=(6, 3)
        )

    def _on_release(self, event):
        x1, y1 = min(self.start_x, event.x), min(self.start_y, event.y)
        x2, y2 = max(self.start_x, event.x), max(self.start_y, event.y)
        if x2 - x1 < 20 or y2 - y1 < 20:
            return
        self.result = (self.winfo_x() + x1, self.winfo_y() + y1,
                       self.winfo_x() + x2, self.winfo_y() + y2)
        self.destroy()


def capture_esi(region=None):
    """捕获屏幕 ESI 数据。region 为 (x1,y1,x2,y2)。返回捕获到的文本。"""
    from PIL import ImageGrab, ImageFilter, ImageOps
    try:
        import pytesseract
        pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
    except Exception:
        pass

    img = ImageGrab.grab(bbox=region)
    # 预处理：放大 + 灰度 + 二值化，提高 OCR 准确率
    w, h = img.size
    img = img.resize((w * 3, h * 3), ImageGrab.Image.LANCZOS if hasattr(ImageGrab.Image, "LANCZOS") else 2)
    img = img.convert("L")
    img = ImageOps.autocontrast(img)
    # 二值化
    threshold = 128
    img = img.point(lambda p: 0 if p < threshold else 255)

    text = pytesseract.image_to_string(img, config="--psm 6 -c tessedit_char_whitelist=0123456789.- ")
    return text.strip()


def parse_esi_numbers(text):
    """从 OCR 结果中提取 ESI 参数字符串，格式如 2.50-40-80-100-20-50-600-4.6"""
    nums = re.findall(r"\d+\.?\d*", text)
    if len(nums) >= 4:
        return "-".join(nums)
    return text


# ============================================================
#  主界面
# ============================================================
class App:
    def __init__(self):
        init_db()
        self.config = load_config()
        self.root = tk.Tk()
        self.root.title("TOF Operation Log")
        self.root.geometry("1250x820")
        self.root.minsize(950, 600)

        self._build_ui()
        self._load_data()
        # Enter 跳转：遍历所有 Entry，绑定 Enter
        self._bind_enter_nav()

    def _bind_enter_nav(self):
        self._input_order = [
            self.ent_date, self.ent_timeperiod, self.ent_name, self.ent_purpose,
            self.ent_solvent, self.cmb_cleaned,
            self.ent_ionsrc, self.ent_sinfo, self.ent_charge, self.ent_speaks,
        ]

        def make_handler(nxt):
            def handler(event):
                nxt.focus_set()
                if isinstance(nxt, ttk.Combobox):
                    try:
                        nxt.event_generate("<Button-1>")
                    except Exception:
                        pass
                elif nxt is self.ent_ionsrc:
                    try:
                        nxt.icursor(tk.END)
                    except Exception:
                        pass
                else:
                    try:
                        nxt.select_range(0, tk.END)
                    except Exception:
                        pass
                return "break"
            return handler

        for i, ent in enumerate(self._input_order):
            nxt = self._input_order[(i + 1) % len(self._input_order)]
            ent.bind("<Return>", make_handler(nxt))

    def _build_ui(self):
        top = ttk.Frame(self.root, padding=12)
        top.pack(fill=tk.X)

        # ---- Row 0: 日期 | 时间段 | 姓名 ----
        r0 = ttk.Frame(top)
        r0.pack(fill=tk.X, pady=(0, 6))

        ttk.Label(r0, text="日期").pack(side=tk.LEFT)
        self.var_date = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        self.ent_date = ttk.Entry(r0, textvariable=self.var_date, width=14)
        self.ent_date.pack(side=tk.LEFT, padx=(4, 16))

        ttk.Label(r0, text="时间段").pack(side=tk.LEFT)
        self.var_timeperiod = tk.StringVar()
        self.ent_timeperiod = ttk.Entry(r0, textvariable=self.var_timeperiod, width=16)
        self.ent_timeperiod.pack(side=tk.LEFT, padx=4)
        ttk.Label(r0, text="例: 6:45-9:00", foreground="gray").pack(side=tk.LEFT, padx=(0, 20))

        ttk.Label(r0, text="姓名").pack(side=tk.LEFT)
        self.var_name = tk.StringVar()
        self.ent_name = ttk.Entry(r0, textvariable=self.var_name, width=16)
        self.ent_name.pack(side=tk.LEFT, padx=4)

        ttk.Button(r0, text="清空表单", command=self._clear_form).pack(side=tk.RIGHT, padx=4)
        ttk.Button(r0, text="保存", command=self._save).pack(side=tk.RIGHT)
        self.btn_cancel = ttk.Button(r0, text="取消编辑", command=self._cancel_edit, state=tk.DISABLED)
        self.btn_cancel.pack(side=tk.RIGHT, padx=4)

        # ---- Row 1: 实验目的 ----
        r1 = ttk.Frame(top)
        r1.pack(fill=tk.X, pady=(0, 6))
        ttk.Label(r1, text="实验目的").pack(side=tk.LEFT)
        self.var_purpose = tk.StringVar()
        self.ent_purpose = ttk.Entry(r1, textvariable=self.var_purpose, width=100)
        self.ent_purpose.pack(side=tk.LEFT, padx=4, fill=tk.X, expand=True)

        # ---- Row 1.5: 溶剂 | 是否清洗干净 ----
        r1b = ttk.Frame(top)
        r1b.pack(fill=tk.X, pady=(0, 6))
        ttk.Label(r1b, text="溶剂").pack(side=tk.LEFT)
        self.var_solvent = tk.StringVar()
        self.ent_solvent = ttk.Entry(r1b, textvariable=self.var_solvent, width=30)
        self.ent_solvent.pack(side=tk.LEFT, padx=4)

        ttk.Label(r1b, text="是否清洗干净").pack(side=tk.LEFT, padx=(20, 0))
        self.var_cleaned = tk.StringVar()
        self.cmb_cleaned = ttk.Combobox(r1b, textvariable=self.var_cleaned, values=["是", "否"], width=6, state="readonly")
        self.cmb_cleaned.pack(side=tk.LEFT, padx=4)

        # ---- Row 2: 离子源 | [捕获ESI] | [设置区域] ----
        r2 = ttk.Frame(top)
        r2.pack(fill=tk.X, pady=(0, 6))
        ttk.Label(r2, text="离子源").pack(side=tk.LEFT)
        self.var_ionsrc = tk.StringVar(value="2.50-40-80-100-20-50-600-4.6")
        self.ent_ionsrc = ttk.Entry(r2, textvariable=self.var_ionsrc, width=70)
        self.ent_ionsrc.pack(side=tk.LEFT, padx=4, fill=tk.X, expand=True)
        ttk.Button(r2, text="设置捕获区域", command=self._set_esi_region).pack(side=tk.LEFT, padx=(12, 4))
        ttk.Button(r2, text="捕获 ESI", command=self._capture_esi).pack(side=tk.LEFT)
        self.lbl_region = ttk.Label(r2, text="", foreground="gray")
        self.lbl_region.pack(side=tk.LEFT, padx=8)

        # 更新区域提示
        esi_region = self.config.get("esi_region")
        if esi_region:
            self.lbl_region.config(text=f"已设置区域: {esi_region}")

        # ---- Row 3: 样品信息 | 电荷 | [计算] ----
        r3 = ttk.Frame(top)
        r3.pack(fill=tk.X, pady=(0, 6))
        ttk.Label(r3, text="样品信息").pack(side=tk.LEFT)
        self.var_sinfo = tk.StringVar()
        self.ent_sinfo = ttk.Entry(r3, textvariable=self.var_sinfo, width=55)
        self.ent_sinfo.pack(side=tk.LEFT, padx=4, fill=tk.X, expand=True)

        ttk.Label(r3, text="电荷").pack(side=tk.LEFT, padx=(12, 0))
        self.var_charge = tk.StringVar()
        self.ent_charge = ttk.Entry(r3, textvariable=self.var_charge, width=8)
        self.ent_charge.pack(side=tk.LEFT, padx=4)
        ttk.Label(r3, text="例: 2+, 1-, +3", foreground="gray").pack(side=tk.LEFT, padx=(0, 8))

        ttk.Button(r3, text="计算样品峰 m/z", command=self._calc_peaks).pack(side=tk.LEFT)

        # ---- Row 4: 样品峰 ----
        r4 = ttk.Frame(top)
        r4.pack(fill=tk.X, pady=(0, 2))
        ttk.Label(r4, text="样品峰").pack(side=tk.LEFT)
        self.var_speaks = tk.StringVar()
        self.ent_speaks = ttk.Entry(r4, textvariable=self.var_speaks, width=100)
        self.ent_speaks.pack(side=tk.LEFT, padx=4, fill=tk.X, expand=True)

        ttk.Separator(self.root, orient=tk.HORIZONTAL).pack(fill=tk.X)

        # ---- 表格工具栏 ----
        bar = ttk.Frame(self.root, padding=(12, 8))
        bar.pack(fill=tk.X)
        ttk.Label(bar, text="搜索:").pack(side=tk.LEFT)
        self.var_search = tk.StringVar()
        self.var_search.trace_add("write", lambda *a: self._apply_filter())
        ttk.Entry(bar, textvariable=self.var_search, width=30).pack(side=tk.LEFT, padx=4)
        ttk.Button(bar, text="导出 CSV", command=self._export_csv).pack(side=tk.RIGHT, padx=4)
        ttk.Button(bar, text="导出 Excel", command=self._export_excel).pack(side=tk.RIGHT, padx=4)
        ttk.Button(bar, text="删除选中", command=self._delete).pack(side=tk.RIGHT, padx=4)
        ttk.Button(bar, text="编辑选中", command=self._edit_selected).pack(side=tk.RIGHT, padx=4)

        # ---- 表格 ----
        tree_frame = ttk.Frame(self.root, padding=(12, 0))
        tree_frame.pack(fill=tk.BOTH, expand=True)

        cols = ("id", "日期", "时间段", "实验目的", "姓名", "离子源信息", "样品信息", "电荷", "样品峰", "溶剂", "是否清洗干净")
        self.tree = ttk.Treeview(tree_frame, columns=cols, show="headings", selectmode="extended")
        widths = [35, 95, 85, 150, 65, 150, 130, 50, 130, 70, 80]
        for c, w in zip(cols, widths):
            self.tree.heading(c, text=c, command=lambda col=c: self._sort(col))
            self.tree.column(c, width=w, minwidth=35, stretch=c not in ("id", "电荷"))

        vsb = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)

        self.tree.bind("<Double-1>", lambda e: self._edit_selected())
        self.root.bind("<Control-s>", lambda e: self._save())
        self.root.bind("<Control-f>", lambda e: self._focus_search())

        # 右键拖拽多选
        self._rclick_start = None
        self.tree.bind("<ButtonPress-3>", self._on_rclick_press)
        self.tree.bind("<B3-Motion>", self._on_rclick_drag)
        self.tree.bind("<ButtonRelease-3>", self._on_rclick_release)

        self.editing_id = None
        self._sort_dir = {}

        self.status = ttk.Label(self.root, text="就绪", relief=tk.SUNKEN, anchor=tk.W, padding=(6, 2))
        self.status.pack(side=tk.BOTTOM, fill=tk.X)

    def _focus_search(self):
        for child in self.root.winfo_children():
            if isinstance(child, ttk.Frame):
                for w in child.winfo_children():
                    if isinstance(w, ttk.Entry):
                        try:
                            if w.cget("textvariable") == str(self.var_search):
                                w.focus_set()
                                return
                        except Exception:
                            pass

    # ---- 右键拖拽多选 ----
    def _on_rclick_press(self, event):
        item = self.tree.identify_row(event.y)
        if item:
            self._rclick_start = item
            self.tree.selection_set(item)
        return "break"

    def _on_rclick_drag(self, event):
        if self._rclick_start is None:
            return "break"
        item = self.tree.identify_row(event.y)
        if not item:
            return "break"
        all_items = self.tree.get_children()
        try:
            start_idx = all_items.index(self._rclick_start)
            cur_idx = all_items.index(item)
        except ValueError:
            return "break"
        lo, hi = min(start_idx, cur_idx), max(start_idx, cur_idx)
        self.tree.selection_set(*all_items[lo:hi + 1])
        return "break"

    def _on_rclick_release(self, event):
        self._rclick_start = None

    # ---- ESI 捕获 ----
    def _set_esi_region(self):
        self.root.iconify()
        self.root.after(300, self._do_region_select)

    def _do_region_select(self):
        sel = RegionSelector(self.root)
        self.root.wait_window(sel)
        self.root.deiconify()
        if sel.result:
            self.config["esi_region"] = list(sel.result)
            save_config(self.config)
            self.lbl_region.config(text=f"已设置区域: {sel.result}")
            self._set_status("ESI 捕获区域已保存")
        else:
            self._set_status("已取消区域设置")

    def _capture_esi(self):
        region = self.config.get("esi_region")
        if not region:
            messagebox.showinfo("提示", "请先点击「设置捕获区域」，框选质谱软件上显示 ESI 数据的区域。")
            return
        try:
            raw = capture_esi(tuple(region))
        except Exception as e:
            messagebox.showerror("捕获失败", f"OCR 出错: {e}")
            return
        if not raw:
            messagebox.showwarning("提示", "未识别到任何内容，请重新设置捕获区域。")
            return
        parsed = parse_esi_numbers(raw)
        self.var_ionsrc.set(parsed)
        self._set_status(f"ESI 已捕获: {parsed}")

    # ---- m/z 计算 ----
    def _calc_peaks(self):
        sinfo = self.var_sinfo.get().strip()
        charge = self.var_charge.get().strip()
        if not sinfo:
            messagebox.showwarning("提示", "请先输入样品信息（分子式）。")
            return
        if not charge:
            messagebox.showwarning("提示", "请先输入电荷信息（如 2+, 1-）。")
            return
        try:
            mz, err = calc_mz(sinfo, charge)
        except ValueError as e:
            messagebox.showerror("分子式解析错误", str(e))
            return
        if err:
            messagebox.showerror("错误", err)
            return
        self.var_speaks.set(f"m/z {mz:.2f}")
        self._set_status(f"已计算: 质量={parse_formula(sinfo)[0]:.4f}, 电荷={charge} → m/z={mz:.2f}")

    # ---- CRUD ----
    def _save(self):
        data = {
            "date": self.var_date.get().strip(),
            "time_period": self.var_timeperiod.get().strip(),
            "purpose": self.var_purpose.get().strip(),
            "name": self.var_name.get().strip(),
            "ion_source": self.var_ionsrc.get().strip(),
            "sample_info": self.var_sinfo.get().strip(),
            "charge": self.var_charge.get().strip(),
            "sample_peaks": self.var_speaks.get().strip(),
            "solvent": self.var_solvent.get().strip(),
            "cleaned": self.var_cleaned.get().strip(),
        }
        if not all(data.values()):
            messagebox.showwarning("提示", "所有字段均为必填，请完整填写。")
            return

        with get_db() as conn:
            if self.editing_id:
                conn.execute(
                    "UPDATE experiments SET date=?, time_period=?, purpose=?, name=?, "
                    "ion_source=?, sample_info=?, charge=?, sample_peaks=?, solvent=?, cleaned=? WHERE id=?",
                    (*data.values(), self.editing_id),
                )
                self._set_status(f"已更新记录 #{self.editing_id}")
            else:
                cur = conn.execute(
                    "INSERT INTO experiments (date, time_period, purpose, name, ion_source, sample_info, charge, sample_peaks, solvent, cleaned) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    tuple(data.values()),
                )
                self._set_status(f"已创建记录 #{cur.lastrowid}")

        self._cancel_edit()
        self._clear_form()
        self._load_data()

    def _load_data(self):
        with get_db() as conn:
            self._all_rows = conn.execute(
                "SELECT * FROM experiments ORDER BY date DESC, time_period DESC"
            ).fetchall()
        self._apply_filter()

    def _apply_filter(self):
        q = self.var_search.get().lower()
        rows = self._all_rows if not q else [
            r for r in self._all_rows
            if any(q in str(v).lower() for v in dict(r).values())
        ]
        self.tree.delete(*self.tree.get_children())
        for r in rows:
            d = dict(r)
            self.tree.insert("", tk.END, iid=str(d["id"]),
                             values=(d["id"], d["date"], d["time_period"], d["purpose"],
                                     d["name"], d["ion_source"], d["sample_info"],
                                     d["charge"], d["sample_peaks"],
                                     d.get("solvent", ""), d.get("cleaned", "")))
        self._auto_fit_columns()
        self._set_status(f"共 {len(rows)} 条记录")

    def _auto_fit_columns(self):
        """根据单元格内容自适应调整表格列宽"""
        import tkinter.font as tkfont
        # ttk.Treeview 不直接支持 cget("font")，从 style 获取或使用默认字体
        try:
            style = ttk.Style()
            font_name = style.lookup(self.tree["style"], "font")
            font = tkfont.Font(font=font_name) if font_name else tkfont.nametofont("TkDefaultFont")
        except Exception:
            font = tkfont.nametofont("TkDefaultFont")
        padding = 28

        cols = ("id", "日期", "时间段", "实验目的", "姓名", "离子源信息", "样品信息", "电荷", "样品峰", "溶剂", "是否清洗干净")
        children = self.tree.get_children()

        for col in cols:
            max_width = font.measure(col) + padding
            col_idx = cols.index(col)
            for child in children:
                values = self.tree.item(child, "values")
                if col_idx < len(values):
                    cell_text = str(values[col_idx])
                    max_width = max(max_width, font.measure(cell_text) + padding)
            max_width = min(max_width, 600)
            min_w = self.tree.column(col)["minwidth"]
            self.tree.column(col, width=max(max_width, min_w))

    def _edit_selected(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("提示", "请先在表格中选中一条记录。")
            return
        rid = int(sel[0])
        row = next((r for r in self._all_rows if r["id"] == rid), None)
        if not row:
            return

        self.editing_id = rid
        self.var_date.set(row["date"])
        self.var_timeperiod.set(row["time_period"])
        self.var_purpose.set(row["purpose"])
        self.var_name.set(row["name"])
        self.var_ionsrc.set(row["ion_source"])
        self.var_sinfo.set(row["sample_info"])
        self.var_charge.set(row["charge"])
        self.var_speaks.set(row["sample_peaks"])
        self.var_solvent.set(row["solvent"] if "solvent" in row.keys() else "")
        self.var_cleaned.set(row["cleaned"] if "cleaned" in row.keys() else "")
        self.btn_cancel.config(state=tk.NORMAL)
        self._set_status(f"正在编辑记录 #{rid}")

    def _cancel_edit(self):
        self.editing_id = None
        self.btn_cancel.config(state=tk.DISABLED)
        self._set_status("就绪")

    def _delete(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("提示", "请先在表格中选中要删除的记录。")
            return
        count = len(sel)
        if not messagebox.askyesno("确认", f"确定要删除选中的 {count} 条记录吗？此操作不可撤销。"):
            return
        ids = [int(s) for s in sel]
        with get_db() as conn:
            conn.executemany("DELETE FROM experiments WHERE id = ?", [(i,) for i in ids])
        if self.editing_id in ids:
            self._cancel_edit()
            self._clear_form()
        self._load_data()
        self._set_status(f"已删除 {count} 条记录")

    def _clear_form(self):
        self.var_date.set(datetime.now().strftime("%Y-%m-%d"))
        self.var_timeperiod.set("")
        self.var_purpose.set("")
        self.var_name.set("")
        self.var_ionsrc.set("2.50-40-80-100-20-50-600-4.6")
        self.var_sinfo.set("")
        self.var_charge.set("")
        self.var_speaks.set("")
        self.var_solvent.set("")
        self.var_cleaned.set("")

    # ---- 排序 ----
    def _sort(self, col):
        rev = self._sort_dir.get(col, False)
        self._sort_dir[col] = not rev
        key_map = {
            "id": lambda r: r["id"],
            "日期": lambda r: r["date"],
            "时间段": lambda r: r["time_period"],
            "实验目的": lambda r: r["purpose"],
            "姓名": lambda r: r["name"],
            "离子源信息": lambda r: r["ion_source"],
            "样品信息": lambda r: r["sample_info"],
            "电荷": lambda r: r["charge"],
            "样品峰": lambda r: r["sample_peaks"],
            "溶剂": lambda r: r.get("solvent", ""),
            "是否清洗干净": lambda r: r.get("cleaned", ""),
        }
        self._all_rows = sorted(self._all_rows, key=key_map.get(col, lambda r: r["id"]), reverse=rev)
        self._apply_filter()

    # ---- 导出 ----
    def _export_csv(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("提示", "请先在表格中选中要导出的记录（可使用鼠标拖拽多选）。")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV 文件", "*.csv")],
            initialfile=f"experiments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
        )
        if not path:
            return

        # 从数据库获取 created_at（Treeview 中不显示此列）
        ids = [int(iid) for iid in sel]
        created_at_map = {}
        try:
            with get_db() as conn:
                placeholders = ",".join("?" * len(ids))
                rows = conn.execute(
                    f"SELECT id, created_at FROM experiments WHERE id IN ({placeholders})",
                    ids,
                ).fetchall()
                created_at_map = {r["id"]: r["created_at"] for r in rows}
        except Exception:
            pass

        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["ID", "日期", "时间段", "实验目的", "姓名", "离子源信息", "样品信息", "电荷", "样品峰", "溶剂", "是否清洗干净", "创建时间"])
            for iid in sel:
                values = list(self.tree.item(iid, "values"))
                rid = int(iid)
                values.append(created_at_map.get(rid, ""))
                w.writerow(values)
        self._set_status(f"已导出 {len(sel)} 条记录到 {os.path.basename(path)}")

    def _export_excel(self):
        """导出为 Excel 文件，自动调整列宽"""
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showwarning("提示", "需要安装 openpyxl 库才能导出 Excel。\n请在命令行运行: pip install openpyxl")
            return

        sel = self.tree.selection()
        if not sel:
            sel = self.tree.get_children()
        if not sel:
            messagebox.showinfo("提示", "表格中没有数据可以导出。")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")],
            initialfile=f"experiments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TOF Operation Log"

        headers = ["ID", "日期", "时间段", "实验目的", "姓名", "离子源信息", "样品信息", "电荷", "样品峰", "溶剂", "是否清洗干净", "创建时间"]
        ws.append(headers)

        # 获取 created_at
        ids = [int(iid) for iid in sel]
        created_at_map = {}
        try:
            with get_db() as conn:
                placeholders = ",".join("?" * len(ids))
                rows = conn.execute(
                    f"SELECT id, created_at FROM experiments WHERE id IN ({placeholders})",
                    ids,
                ).fetchall()
                created_at_map = {r["id"]: r["created_at"] for r in rows}
        except Exception:
            pass

        for iid in sel:
            values = list(self.tree.item(iid, "values"))
            rid = int(iid)
            values.append(created_at_map.get(rid, ""))
            ws.append(values)

        # 自适应列宽：中文约占 2 个字符宽度
        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_len = 0
            for cell in col_cells:
                if cell.value is not None:
                    cell_str = str(cell.value)
                    char_len = 0
                    for ch in cell_str:
                        if ord(ch) > 127:
                            char_len += 2
                        else:
                            char_len += 1
                    max_len = max(max_len, char_len)
            adjusted = min(max_len + 4, 80)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted

        # 表头加粗
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)

        wb.save(path)
        self._set_status(f"已导出 {len(sel)} 条记录到 {os.path.basename(path)}")

    def _set_status(self, msg):
        self.status.config(text=msg)

    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    App().run()
